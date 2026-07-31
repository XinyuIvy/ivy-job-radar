from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(
    "/workspace/scratch/a463281d9ca1/outputs/job_company_expansion/"
    "biostatistics_job_application_tracker_expanded.xlsx"
)
OUTPUT = Path(__file__).resolve().parents[1] / "app" / "company-pool.json"


def clean(value: object) -> str | int:
    if value is None:
        return ""
    return value if isinstance(value, int) else str(value).strip()


workbook = load_workbook(SOURCE, read_only=False, data_only=True)
sheet = workbook["目标公司与职位"]
companies: list[dict[str, str | int]] = []

for row in sheet.iter_rows(min_row=3, values_only=True):
    if not row[5]:
        continue
    companies.append(
        {
            "rank": clean(row[0]),
            "region": clean(row[1]),
            "priority": clean(row[2]),
            "track": clean(row[3]),
            "fit": clean(row[4]),
            "company": clean(row[5]),
            "companyType": clean(row[6]),
            "keywords": clean(row[7]),
            "reason": clean(row[8]),
            "dataTypes": clean(row[9]),
            "strategy": clean(row[10]),
            "source": clean(row[11]),
        }
    )

OUTPUT.write_text(
    json.dumps(companies, ensure_ascii=False, indent=2) + "\n",
    encoding="utf-8",
)
print(f"Wrote {len(companies)} companies to {OUTPUT}")
